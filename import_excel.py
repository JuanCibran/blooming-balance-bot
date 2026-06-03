"""
Importa todos los datos del Excel viejo al Google Sheet.
Corre una sola vez: python import_excel.py
"""

import json
import openpyxl
import gspread
from google.oauth2 import service_account
from config import GOOGLE_APPLICATION_CREDENTIALS, GOOGLE_CREDENTIALS_JSON, SPREADSHEET_ID

EXCEL_PATH = "/Users/juancruzcibran/Downloads/blooming essie control 2.xlsx"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]


def get_client():
    if GOOGLE_CREDENTIALS_JSON:
        info = json.loads(GOOGLE_CREDENTIALS_JSON)
        creds = service_account.Credentials.from_service_account_info(info, scopes=SCOPES)
    else:
        creds = service_account.Credentials.from_service_account_file(
            GOOGLE_APPLICATION_CREDENTIALS, scopes=SCOPES
        )
    return gspread.authorize(creds)


def format_date(value):
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y")
    return str(value)


def import_registro_diario(wb_excel, sheet_gsheets):
    ws = wb_excel["📋 Registro Diario"]

    rows_to_insert = []
    for row in ws.iter_rows(min_row=4, max_row=56, values_only=True):
        # Saltear filas completamente vacías
        if not any(v is not None for v in row):
            continue
        # Saltear filas de totales/resumen
        tipo = str(row[1]).strip().lower() if row[1] else ""
        if tipo not in ("venta", "gasto", "reinversión", "reinversion"):
            continue

        rows_to_insert.append([
            format_date(row[0]),          # Fecha
            str(row[1]).capitalize() if row[1] else "",  # Tipo
            row[2] or "",                 # Categoría
            row[3] or "",                 # Descripción
            row[4] or "",                 # Producto/SKU
            row[5] or "",                 # n orden
            row[6] or "",                 # Precio Unit.
            row[7] or "",                 # Total ($)
            str(row[8]).upper() if row[8] else "",  # Medio de pago
            row[9] or "",                 # Notas
        ])

    if rows_to_insert:
        sheet_gsheets.update(
            values=rows_to_insert,
            range_name=f"A4:J{3 + len(rows_to_insert)}",
            value_input_option="USER_ENTERED"
        )
        print(f"  ✅ {len(rows_to_insert)} filas importadas al Registro Diario")
    else:
        print("  ⚠️  No se encontraron filas en el Registro Diario")


def import_stock(wb_excel, sheet_gsheets):
    ws = wb_excel["📦 Stock"]

    rows_to_insert = []
    for row in ws.iter_rows(min_row=3, max_row=28, values_only=True):
        if not row[0]:  # SKU vacío = fila vacía
            continue
        rows_to_insert.append([
            row[0] or "",   # SKU
            row[1] or "",   # Producto
            row[2] or "",   # Categoría
            row[3] or "",   # Talle/Color
            row[4] or 0,    # Stock inicial
            row[5] or 0,    # Entradas
            row[6] or 0,    # Salidas
            "",             # Stock actual (tiene fórmula en el Sheet)
            row[8] or "",   # Precio costo
            row[9] or "",   # Precio venta
        ])

    if rows_to_insert:
        # Reemplazar filas desde A3
        sheet_gsheets.update(
            values=rows_to_insert,
            range_name=f"A3:J{2 + len(rows_to_insert)}",
            value_input_option="USER_ENTERED"
        )
        print(f"  ✅ {len(rows_to_insert)} productos importados al Stock")
    else:
        print("  ⚠️  No se encontraron productos en el Stock")


def main():
    print("📂 Leyendo Excel...")
    wb_excel = openpyxl.load_workbook(EXCEL_PATH)

    print("🔗 Conectando al Google Sheet...")
    client = get_client()
    wb_sheets = client.open_by_key(SPREADSHEET_ID)

    print("\n📋 Importando Registro Diario...")
    import_registro_diario(wb_excel, wb_sheets.worksheet("📋 Registro Diario"))

    print("\n📦 Importando Stock...")
    import_stock(wb_excel, wb_sheets.worksheet("📦 Stock"))

    print("\n🎉 Importación completa.")
    print(f"🔗 Abrí el Sheet: https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}")


if __name__ == "__main__":
    main()
